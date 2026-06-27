import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import datetime

def create_dashboard():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Global styles
    font_segoe = Font(name='Segoe UI')
    font_segoe_bold = Font(name='Segoe UI', bold=True)
    font_segoe_title = Font(name='Segoe UI', size=16, bold=True, color="1F4E78")
    currency_style = NamedStyle(name="currency", number_format='#,##0" ₫"', font=Font(name='Segoe UI'))
    percent_style = NamedStyle(name="percent", number_format='0.0%', font=Font(name='Segoe UI'))
    
    if "currency" not in wb.named_styles:
        wb.add_named_style(currency_style)
    if "percent" not in wb.named_styles:
        wb.add_named_style(percent_style)
        
    def apply_font(ws):
        for row in ws.iter_rows():
            for cell in row:
                if not cell.font or cell.font.name != 'Segoe UI':
                    if cell.font:
                        cell.font = Font(name='Segoe UI', bold=cell.font.bold, size=cell.font.size, color=cell.font.color)
                    else:
                        cell.font = font_segoe

    # ==========================================
    # 1. Sheet 1: Data_NhapLieu
    # ==========================================
    ws_data = wb.active
    ws_data.title = "Data_NhapLieu"
    
    headers = [
        "Date", "Invoice", "Patient", "Consultation fees", 
        "Pharmacy sales", "Other revenue", "DETAILS BY PAYMENT",
        "Total Amount", "Payment time"
    ]
    
    header_font = Font(name='Segoe UI', bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    sample_data = [
        [datetime.date(2026, 6, 26), "INV001", "Nguyen Van A", 150000, 50000, 0, "Cash", 200000, "08:30:00"],
        [datetime.date(2026, 6, 26), "INV002", "Tran Thi B", 0, 150000, 0, "Bank Transfer / Credit Card", 150000, "09:15:00"],
        [datetime.date(2026, 6, 27), "INV003", "Le Van C", 150000, 100000, 100000, "Cash", 350000, "10:00:00"],
    ]
    
    # Format up to 1000 rows so user doesn't need to manually format
    for row_idx in range(2, 1001):
        # Format Date for column A (1)
        cell_date = ws_data.cell(row=row_idx, column=1)
        cell_date.number_format = 'yyyy-mm-dd'
        # Format Currency for columns D, E, F, H (4, 5, 6, 8)
        for col_idx in [4, 5, 6, 8]:
            cell_currency = ws_data.cell(row=row_idx, column=col_idx)
            cell_currency.number_format = '#,##0" ₫"'
            cell_currency.font = font_segoe
        # Alignments
        alignments = {
            1: Alignment(horizontal="center", vertical="center"), # Date
            2: Alignment(horizontal="center", vertical="center"), # Invoice
            3: Alignment(horizontal="left", vertical="center"),   # Patient
            4: Alignment(horizontal="right", vertical="center"),  # Consultation
            5: Alignment(horizontal="right", vertical="center"),  # Pharmacy
            6: Alignment(horizontal="right", vertical="center"),  # Other
            7: Alignment(horizontal="center", vertical="center"), # Payment Method
            8: Alignment(horizontal="right", vertical="center"),  # Total
            9: Alignment(horizontal="center", vertical="center"), # Time
        }
        # Format standard Font for all columns
        for col_idx in range(1, 10):
            cell = ws_data.cell(row=row_idx, column=col_idx)
            cell.font = font_segoe
            cell.alignment = alignments[col_idx]
    
    # Fill in sample data
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=val)
                
    for col in range(1, 10):
        ws_data.column_dimensions[get_column_letter(col)].width = 20
        
    # --- ADD DATA VALIDATION (DROPDOWNS) ---
    dv_payment = DataValidation(type="list", formula1='"Cash,Bank Transfer / Credit Card"', allow_blank=True)
    ws_data.add_data_validation(dv_payment)
    dv_payment.add("G2:G1000")

    # ==========================================
    # 2. Sheet 2: Daily_Dashboard
    # ==========================================
    ws_daily = wb.create_sheet("Daily_Dashboard")
    ws_daily.sheet_view.showGridLines = True
    
    # Widths
    ws_daily.column_dimensions['B'].width = 10
    ws_daily.column_dimensions['C'].width = 45
    ws_daily.column_dimensions['D'].width = 25
    ws_daily.column_dimensions['E'].width = 25
    
    ws_daily.merge_cells('C2:E2')
    ws_daily['C2'] = "CLINIC DAILY REVENUE REPORT"
    ws_daily['C2'].font = font_segoe_title
    ws_daily['C2'].alignment = Alignment(horizontal="center")
    
    ws_daily['C3'] = "Ngày báo cáo (YYYY-MM-DD):"
    ws_daily['C3'].font = font_segoe_bold
    ws_daily['C3'].alignment = Alignment(horizontal="right")
    ws_daily['D3'] = "2026-06-26"
    ws_daily['D3'].font = font_segoe_bold
    
    ws_daily['C5'] = "💰 TOTAL ACTUAL REVENUE:"
    ws_daily['C5'].font = Font(name='Segoe UI', bold=True, size=12)
    ws_daily['E5'] = "=SUMIF(Data_NhapLieu!A:A, D3, Data_NhapLieu!H:H)"
    ws_daily['E5'].font = Font(name='Segoe UI', bold=True, size=12)
    ws_daily['E5'].style = "currency"
    
    ws_daily['C7'] = "📍 DETAILS BY REVENUE SOURCE (System):"
    ws_daily['C7'].font = Font(name='Segoe UI', bold=True, size=12)
    
    ws_daily['C8'] = "♦ Consultation fees:"
    ws_daily['E8'] = "=SUMIFS(Data_NhapLieu!D:D, Data_NhapLieu!A:A, D3)"
    ws_daily['E8'].style = "currency"
    
    ws_daily['C9'] = "♦ Pharmacy sales:"
    ws_daily['E9'] = "=SUMIFS(Data_NhapLieu!E:E, Data_NhapLieu!A:A, D3)"
    ws_daily['E9'].style = "currency"
    
    ws_daily['C10'] = "♦ Other revenue (Procedures/Tests):"
    ws_daily['E10'] = "=SUMIFS(Data_NhapLieu!F:F, Data_NhapLieu!A:A, D3)"
    ws_daily['E10'].style = "currency"
    
    ws_daily['C12'] = "💳 DETAILS BY PAYMENT METHOD:"
    ws_daily['C12'].font = Font(name='Segoe UI', bold=True, size=12)
    
    ws_daily['C13'] = "▪ Cash:"
    ws_daily['E13'] = '=SUMIFS(Data_NhapLieu!H:H, Data_NhapLieu!A:A, D3, Data_NhapLieu!G:G, "Cash")'
    ws_daily['E13'].style = "currency"
    
    ws_daily['C14'] = "▪ Bank Transfer / Credit Card:"
    ws_daily['E14'] = '=SUMIFS(Data_NhapLieu!H:H, Data_NhapLieu!A:A, D3, Data_NhapLieu!G:G, "Bank Transfer / Credit Card")'
    ws_daily['E14'].style = "currency"
    
    ws_daily['C16'] = "👥 PATIENT STATISTICS:"
    ws_daily['C16'].font = Font(name='Segoe UI', bold=True, size=12)
    
    ws_daily['C17'] = "▪ Total Patients (Tổng số ca khám):"
    ws_daily['E17'] = '=COUNTIF(Data_NhapLieu!A:A, D3)'
    
    ws_daily['C18'] = "▪ New Patients (Bệnh nhân mới):"
    ws_daily['E18'] = '=COUNTIFS(Data_NhapLieu!A:A, D3, Data_NhapLieu!D:D, "Mới")'
    
    apply_font(ws_daily)

    # ==========================================
    # 3. Sheet 3: Monthly_Dashboard
    # ==========================================
    ws_monthly = wb.create_sheet("Monthly_Dashboard")
    ws_monthly.sheet_view.showGridLines = True
    
    # Widths
    ws_monthly.column_dimensions['B'].width = 10
    ws_monthly.column_dimensions['C'].width = 45
    ws_monthly.column_dimensions['D'].width = 15
    ws_monthly.column_dimensions['E'].width = 25
    ws_monthly.column_dimensions['F'].width = 15
    
    ws_monthly.merge_cells('C2:E2')
    ws_monthly['C2'] = "MONTHLY REVENUE SUMMARY REPORT"
    ws_monthly['C2'].font = font_segoe_title
    ws_monthly['C2'].alignment = Alignment(horizontal="center")
    
    ws_monthly['C3'] = "Month / Year:"
    ws_monthly['C3'].font = font_segoe_bold
    ws_monthly['C3'].alignment = Alignment(horizontal="right")
    ws_monthly['D3'] = 6
    ws_monthly['D3'].font = font_segoe_bold
    ws_monthly['E3'] = 2026
    ws_monthly['E3'].font = font_segoe_bold
    
    ws_monthly['C5'] = "💰 TOTAL ACTUAL REVENUE:"
    ws_monthly['C5'].font = Font(name='Segoe UI', bold=True, size=12)
    ws_monthly['E5'] = '=SUMIFS(Data_NhapLieu!H:H, Data_NhapLieu!A:A, ">="&DATE(E3,D3,1), Data_NhapLieu!A:A, "<="&EOMONTH(DATE(E3,D3,1),0))'
    ws_monthly['E5'].font = Font(name='Segoe UI', bold=True, size=12)
    ws_monthly['E5'].style = "currency"
    
    ws_monthly['C7'] = "📍 DETAILED REVENUE STRUCTURE:"
    ws_monthly['C7'].font = Font(name='Segoe UI', bold=True, size=12)
    
    ws_monthly['C8'] = "♦ Consultation fees:"
    ws_monthly['E8'] = '=SUMIFS(Data_NhapLieu!D:D, Data_NhapLieu!A:A, ">="&DATE(E3,D3,1), Data_NhapLieu!A:A, "<="&EOMONTH(DATE(E3,D3,1),0))'
    ws_monthly['E8'].style = "currency"
    ws_monthly['F8'] = "=IF(E5>0, E8/E5, 0)"
    ws_monthly['F8'].style = "percent"
    
    ws_monthly['C9'] = "♦ Pharmacy sales:"
    ws_monthly['E9'] = '=SUMIFS(Data_NhapLieu!E:E, Data_NhapLieu!A:A, ">="&DATE(E3,D3,1), Data_NhapLieu!A:A, "<="&EOMONTH(DATE(E3,D3,1),0))'
    ws_monthly['E9'].style = "currency"
    ws_monthly['F9'] = "=IF(E5>0, E9/E5, 0)"
    ws_monthly['F9'].style = "percent"
    
    ws_monthly['C10'] = "♦ Other services revenue:"
    ws_monthly['E10'] = '=SUMIFS(Data_NhapLieu!F:F, Data_NhapLieu!A:A, ">="&DATE(E3,D3,1), Data_NhapLieu!A:A, "<="&EOMONTH(DATE(E3,D3,1),0))'
    ws_monthly['E10'].style = "currency"
    ws_monthly['F10'] = "=IF(E5>0, E10/E5, 0)"
    ws_monthly['F10'].style = "percent"
    
    ws_monthly['C12'] = "👥 PATIENT STATISTICS:"
    ws_monthly['C12'].font = Font(name='Segoe UI', bold=True, size=12)
    
    ws_monthly['C13'] = "▪ Total patient visits this month:"
    ws_monthly['E13'] = '=COUNTIFS(Data_NhapLieu!A:A, ">="&DATE(E3,D3,1), Data_NhapLieu!A:A, "<="&EOMONTH(DATE(E3,D3,1),0))'
    
    ws_monthly['C14'] = "▪ Average patients/day:"
    ws_monthly['E14'] = '=COUNTIFS(Data_NhapLieu!A:A, ">="&DATE(E3,D3,1), Data_NhapLieu!A:A, "<="&EOMONTH(DATE(E3,D3,1),0)) / DAY(EOMONTH(DATE(E3,D3,1),0))'
    # Format average to 1 decimal place
    avg_style = NamedStyle(name="average", number_format='0.0', font=Font(name='Segoe UI'))
    if "average" not in wb.named_styles:
        wb.add_named_style(avg_style)
    ws_monthly['E14'].style = "average"
    
    apply_font(ws_monthly)
    
    wb.save("Clinic_Revenue_Reports_V7.xlsx")
    print("Excel template updated successfully with formatted columns and Dropdown list!")

if __name__ == "__main__":
    create_dashboard()
